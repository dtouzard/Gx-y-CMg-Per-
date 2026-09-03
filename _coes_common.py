# -*- coding: utf-8 -*-
"""
Funciones compartidas entre CMg_lookup_Peru.py, Generacion_lookup_Peru.py y
Cruce_Ingresos_Peru.py.

Se centralizaron aca porque antes vivian duplicadas (copiadas literal) en
cada script, y esa duplicacion ya causo bugs reales dos veces (ver
CLAUDE.md, decision tecnica #3): al corregir la logica en una copia, la
otra quedaba con el bug viejo hasta que alguien se acordaba de replicar el
fix a mano. Con esto, un fix a _fecha_operativa (por ejemplo) se aplica una
sola vez y los tres scripts lo heredan al importar.

No se movio nada que sea especifico de un solo script (ej. dias_rango, que
solo usa Generacion_lookup_Peru.py).
"""

import os, re
import pandas as pd


def _limpia_numero(serie):
    """Convierte a numero admitiendo tanto '1234.56' (CSV de Generacion)
    como '1234,56' / '2 056,628' (formato COES con coma decimal y espacio
    de miles, ver CLAUDE.md decision tecnica #7). Los valores que no se
    puedan convertir quedan en NaN (pd.to_numeric con errors="coerce") -
    quien llama a esta funcion es responsable de avisar si el numero de
    NaN resultante es sospechosamente alto, en vez de descartarlos en
    silencio."""
    s = serie.astype(str).str.strip()
    tiene_coma = s.str.contains(",", regex=False)
    # Si el valor tiene coma Y punto, asumimos punto=miles, coma=decimal.
    tiene_ambos = tiene_coma & s.str.contains(".", regex=False)
    s = s.where(~tiene_ambos, s.str.replace(".", "", regex=False))
    s = s.where(~tiene_coma, s.str.replace(",", ".", regex=False))
    return pd.to_numeric(s, errors="coerce")


def fmt_dur(seg):
    if seg < 60:
        return f"{seg:.1f} s"
    m, s = divmod(seg, 60)
    return f"{int(m)} min {s:.0f} s"


def siguiente_version(carpeta, hoy, base):
    import glob
    maxv = 0
    for f in glob.glob(os.path.join(carpeta, f"{hoy} - {base} v*.xlsx")):
        m = re.search(r"v(\d+)\.xlsx$", os.path.basename(f))
        if m:
            maxv = max(maxv, int(m.group(1)))
    return maxv + 1


def periodos_rango(anio_ini, mes_ini, anio_fin, mes_fin):
    """Lista de (anio, mes) entre los limites, inclusive."""
    out = []
    a, m = anio_ini, mes_ini
    while (a, m) <= (anio_fin, mes_fin):
        out.append((a, m))
        m += 1
        if m > 12:
            m = 1
            a += 1
    return out


def _fecha_operativa(fecha_hora):
    """La hora "00:00" de un dia representa el intervalo 23:00-24:00 del
    dia ANTERIOR (se arma con el punto que cierra a las 23:45 de ese dia
    anterior), no el inicio del dia que indica su propio timestamp. Si se
    sacan anio/mes/dia directo del timestamp, esa hora queda mal atribuida
    al dia SIGUIENTE. Por eso, solo para esos puntos de medianoche, se
    resta 1 dia antes de leer anio/mes/dia (la resta con Timedelta hace
    que tambien se corrija bien cuando la medianoche cae justo en cambio
    de mes o de anio, ej. 01/02 00:00 -> 31/01)."""
    es_medianoche = fecha_hora.dt.hour == 0
    return fecha_hora - pd.to_timedelta(es_medianoche.astype(int), unit="D")


def _nombre_hoja(nombre, usados):
    """Sanitiza un nombre de barra/central para usarlo como nombre de hoja
    de Excel (max 31 caracteres, sin : \\ / ? * [ ]) y evita choques si dos
    nombres truncados quedan iguales."""
    limpio = re.sub(r'[:\\/?*\[\]]', "_", nombre)[:31]
    final = limpio
    i = 2
    while final in usados:
        sufijo = f"_{i}"
        final = limpio[: 31 - len(sufijo)] + sufijo
        i += 1
    usados.add(final)
    return final


def leer_parametros_comunes(ruta=None):
    """Lee `0. Inputs/Parametros_Comunes.xlsx` (hoja compartida por los 4
    scripts, generada por `0. Inputs/Generar_Parametros_Comunes.py`),
    pensada para que gente no tecnica -incluidos gerentes- pueda cambiar
    periodo/barras/empresas/ajustes sin tocar codigo Python. Antes cada
    script tenia su propia fecha inicio/fin en su seccion PARAMETROS, y
    eso ya causo un bug real (Cruce no encontraba cache porque CMg y
    Generacion habian quedado con periodos distintos) - de ahi centralizar
    esto en un solo lugar.

    Frena (SystemExit) con un mensaje claro si falta el archivo o si algun
    valor no tiene el formato esperado, en vez de fallar en silencio mas
    adelante por un typo en una celda.

    Devuelve un dict:
      anio_ini, mes_ini, anio_fin, mes_fin (int)
      barras (list[str], vacia = TODAS)
      empresas (list[str], vacia = TODAS)
      tipos_generacion (list[str], vacia = TODOS)
      agrupar_por_central, convertir_a_usd (bool)
      fecha_base_real (str "AAAA-MM")
      central_categoria_label (str: "TODOS" / "COES" / "GENERACION RER" -
        el llamador la mapea al codigo numerico de Generacion_lookup_Peru.py,
        via CENTRAL_CATEGORIAS, para no hardcodear ese mapeo aca)
      cruce_objetivo (dict {central: {"barra":..., "potencia_central":...}},
        para CRUCE_OBJETIVO de Cruce_Ingresos_Peru.py - lee la hoja "Cruce",
        tabla de 3 columnas Central/Barra/Potencia)."""
    import openpyxl

    if ruta is None:
        ruta = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                             "0. Inputs", "Parametros_Comunes.xlsx")
    if not os.path.exists(ruta):
        raise SystemExit(
            f"No se encontro {ruta}. Corre '0. Inputs/Generar_Parametros_Comunes.py' "
            f"una vez para crearlo (o revisa que no lo hayan movido/renombrado)."
        )
    import warnings
    with warnings.catch_warnings():
        # openpyxl avisa que no soporta la extension de validacion de datos
        # de Excel (los dropdowns) al LEER el archivo - no afecta la lectura
        # de los valores de las celdas (que es todo lo que hacemos aca, este
        # modulo nunca vuelve a guardar el archivo), es puro ruido en consola.
        warnings.filterwarnings("ignore", message="Data Validation extension is not supported")
        wb = openpyxl.load_workbook(ruta, data_only=True)

    def _valor(hoja, etiqueta):
        ws = wb[hoja]
        for fila in ws.iter_rows(min_row=2, values_only=True):
            if fila[0] == etiqueta:
                return fila[1]
        raise SystemExit(f"No se encontro la fila '{etiqueta}' en la hoja '{hoja}' de {ruta}.")

    def _lista(hoja):
        ws = wb[hoja]
        out = []
        for fila in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            v = fila[0]
            if v is not None and str(v).strip():
                out.append(str(v).strip())
        return out

    def _entero(hoja, etiqueta):
        v = _valor(hoja, etiqueta)
        try:
            return int(v)
        except (TypeError, ValueError):
            raise SystemExit(f"'{etiqueta}' en la hoja '{hoja}' deberia ser un numero, vino: {v!r}")

    def _si_no(hoja, etiqueta):
        v = str(_valor(hoja, etiqueta)).strip().lower()
        if v in ("sí", "si", "s", "yes", "true"):
            return True
        if v in ("no", "n", "false"):
            return False
        raise SystemExit(f"'{etiqueta}' en la hoja '{hoja}' deberia ser Sí/No, vino: {v!r}")

    anio_ini = _entero("Periodo", "Año inicio")
    mes_ini = _entero("Periodo", "Mes inicio (1-12)")
    anio_fin = _entero("Periodo", "Año fin")
    mes_fin = _entero("Periodo", "Mes fin (1-12)")
    if not (1 <= mes_ini <= 12 and 1 <= mes_fin <= 12):
        raise SystemExit("El mes de inicio/fin en la hoja 'Periodo' tiene que estar entre 1 y 12.")
    if (anio_ini, mes_ini) > (anio_fin, mes_fin):
        raise SystemExit(f"El periodo en la hoja 'Periodo' esta al reves: inicio "
                          f"({anio_ini}-{mes_ini:02d}) es posterior al fin ({anio_fin}-{mes_fin:02d}).")

    # Hoja "Cruce": tabla de 3 columnas (Central | Barra | Potencia), a
    # diferencia de las demas hojas no es una lista de un solo valor por
    # fila sino un mapeo central->{barra, potencia_central} para
    # CRUCE_OBJETIVO de Cruce_Ingresos_Peru.py. Una fila cuenta si tiene
    # las 3 columnas completas; si solo tiene alguna, se avisa (fila a
    # medio completar, probable error de tipeo/copy-paste).
    ws_cruce = wb["Cruce"]
    cruce_objetivo = {}
    filas_incompletas = []
    for fila in ws_cruce.iter_rows(min_row=2, max_col=3, values_only=True):
        central, barra, potencia = (str(v).strip() if v is not None else "" for v in fila)
        if not central and not barra and not potencia:
            continue
        if not (central and barra and potencia):
            filas_incompletas.append((central, barra, potencia))
            continue
        cruce_objetivo[central] = {"barra": barra, "potencia_central": potencia}
    if filas_incompletas:
        raise SystemExit(
            f"La hoja 'Cruce' tiene {len(filas_incompletas)} fila(s) a medio completar (les "
            f"falta Central, Barra o Potencia): {filas_incompletas[:5]}. Completá las 3 "
            f"columnas de esa fila o borrala entera."
        )

    return {
        "anio_ini": anio_ini, "mes_ini": mes_ini, "anio_fin": anio_fin, "mes_fin": mes_fin,
        "barras": _lista("Barras"),
        "empresas": _lista("Empresas"),
        "tipos_generacion": _lista("Tipos de generación"),
        "agrupar_por_central": _si_no("Otros ajustes", "Agrupar generación por central (Sí/No)"),
        "convertir_a_usd": _si_no("Otros ajustes", "Convertir CMg a USD (Sí/No)"),
        "fecha_base_real": str(_valor("Otros ajustes", "Fecha base USD real (AAAA-MM)")).strip(),
        "central_categoria_label": str(_valor("Otros ajustes", "Categoría de central (Generación)")).strip(),
        "cruce_objetivo": cruce_objetivo,
    }


def escribir_hoja_con_bloques(writer, sheet_name, bloques):
    """Escribe varias tablas apiladas en UNA misma hoja, cada una con un
    titulo en negrita arriba y una fila en blanco antes de la siguiente
    (pedido explicito: agregados por barra/central en una sola hoja con
    titulos, no una hoja por agregado)."""
    from openpyxl.styles import Font
    fila0 = 0  # proxima fila libre (0-indexada)
    for titulo, df in bloques:
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=fila0 + 1)
        ws = writer.sheets[sheet_name]
        ws.cell(row=fila0 + 1, column=1, value=titulo).font = Font(bold=True)
        fila0 += len(df) + 3  # titulo(1) + encabezado(1) + datos + 1 fila en blanco
